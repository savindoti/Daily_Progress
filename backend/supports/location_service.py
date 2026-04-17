"""Service to read and cache location data from Excel file."""
import os
from pathlib import Path
from openpyxl import load_workbook

_LOCATION_DATA = None


def load_location_data():
    """Load location data from Excel file."""
    global _LOCATION_DATA
    
    if _LOCATION_DATA is not None:
        return _LOCATION_DATA
    
    # Try to find the Excel file
    excel_files = [
        Path(__file__).parent.parent / "Book1.xlsx",
        Path(__file__).parent.parent / "locations.xlsx",
        Path.cwd() / "Book1.xlsx",
    ]
    
    excel_path = None
    for path in excel_files:
        if path.exists():
            excel_path = path
            break
    
    if not excel_path:
        # Return empty structure if file not found
        return {"provinces": [], "districts": {}, "municipalities": {}}
    
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        provinces = []
        districts = {}  # province -> list of districts
        municipalities = {}  # district -> list of municipals
        
        # Skip header row
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx == 1:  # Skip header
                continue
            
            if not row or len(row) < 3:
                continue
            
            province = row[0]
            district = row[1]
            municipality = row[2]
            
            if not province or not district or not municipality:
                continue
            
            province = str(province).strip()
            district = str(district).strip()
            municipality = str(municipality).strip()
            
            # Build provinces list (unique)
            if province not in provinces:
                provinces.append(province)
            
            # Build districts dict
            if province not in districts:
                districts[province] = []
            if district not in districts[province]:
                districts[province].append(district)
            
            # Build municipalities dict
            if district not in municipalities:
                municipalities[district] = []
            if municipality not in municipalities[district]:
                municipalities[district].append(municipality)
        
        _LOCATION_DATA = {
            "provinces": sorted(provinces),
            "districts": {k: sorted(v) for k, v in districts.items()},
            "municipalities": {k: sorted(v) for k, v in municipalities.items()},
        }
        
        return _LOCATION_DATA
    
    except Exception as e:
        print(f"Error loading location data: {e}")
        return {"provinces": [], "districts": {}, "municipalities": {}}


def get_locations():
    """Get all location data."""
    return load_location_data()


def get_provinces():
    """Get list of all provinces."""
    return load_location_data()["provinces"]


def get_districts_for_province(province):
    """Get districts for a specific province."""
    data = load_location_data()
    return data.get("districts", {}).get(province, [])


def get_municipalities_for_district(district):
    """Get municipalities for a specific district."""
    data = load_location_data()
    return data.get("municipalities", {}).get(district, [])
